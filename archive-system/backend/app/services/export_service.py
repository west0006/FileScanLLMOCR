"""
通用 Excel 导出工具

用法:
  from app.services.export import export_to_excel
  path = export_to_excel("搜索结果", [{"档案编号":"x","题名":"y"}], ["档案编号","题名"])
"""

import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_to_excel(title: str, rows: list[dict], columns: list[str], output_dir: str = "/tmp") -> str:
    """生成 Excel 文件，返回文件路径"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # sheet 名最长 31 字符

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True, color="1E2130")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # 副标题（导出时间）
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    time_cell = ws.cell(row=2, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   共 {len(rows)} 条")
    time_cell.font = Font(size=10, color="9CA3AF")
    time_cell.alignment = Alignment(horizontal="center")

    # 表头
    header_fill = PatternFill(start_color="F5F6FA", end_color="F5F6FA", fill_type="solid")
    header_font = Font(size=11, bold=True, color="6B7280")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 24

    # 数据行
    data_font = Font(size=11, color="1E2130")
    for row_idx, row in enumerate(rows, 5):
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

    # 列宽自适应
    for col_idx in range(1, len(columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(4, 5 + min(len(rows), 100))
        )
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # 保存
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{title}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def _register_cjk_font(pdf) -> None:
    """注册中文字体（跨平台查找）"""
    candidates = [
        # Windows
        "C:/Windows/Fonts/simhei.ttf",      # 黑体
        "C:/Windows/Fonts/msyh.ttc",        # 微软雅黑
        "C:/Windows/Fonts/simsun.ttc",      # 宋体
        # Linux
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
        # macOS
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
    ]
    import os as _os
    for fp in candidates:
        if _os.path.exists(fp):
            try:
                pdf.add_font("CJK", "", fp)
                pdf.add_font("CJK", "B", fp)
                return "CJK"
            except Exception:
                continue
    return None


def export_to_pdf(title: str, rows: list[dict], columns: list[str], output_dir: str = "/tmp") -> str:
    """生成 PDF 文件，返回文件路径。"""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    font = _register_cjk_font(pdf) or "Helvetica"

    # 标题
    pdf.set_font(font, "B", 14)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    # 导出时间
    pdf.set_font(font, "", 9)
    pdf.set_text_color(128, 128, 128)
    pdf.cell(0, 6, f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   共 {len(rows)} 条",
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(6)

    # 表头
    pdf.set_fill_color(245, 246, 250)
    pdf.set_text_color(100, 100, 100)
    pdf.set_font(font, "B", 9)
    col_w = (pdf.w - 20) / len(columns)
    for col in columns:
        pdf.cell(col_w, 8, col[:20], border=1, fill=True, align="L")
    pdf.ln()

    # 数据行
    pdf.set_text_color(30, 30, 30)
    pdf.set_font(font, "", 8)
    for row in rows:
        for col in columns:
            val = str(row.get(col, ""))[:40]
            if isinstance(row.get(col), list):
                val = ", ".join(str(v) for v in row.get(col, []))[:40]
            pdf.cell(col_w, 7, val, border=1, align="L")
        pdf.ln()

    os.makedirs(output_dir, exist_ok=True)
    filename = f"{title}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


def export_to_csv(title: str, rows: list[dict], columns: list[str], output_dir: str = "/tmp") -> str:
    """生成 CSV 文件（UTF-8 BOM，Excel 可直接打开），返回文件路径"""
    import csv as _csv
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{title}_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    filepath = os.path.join(output_dir, filename)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = _csv.writer(f)
        writer.writerow(columns)
        for row in rows:
            writer.writerow([
                ", ".join(str(v) for v in row.get(col, [])) if isinstance(row.get(col, ""), list) else row.get(col, "")
                for col in columns
            ])
    return filepath
